#!/usr/bin/env python3
"""
Genera fixtures Excel para los tests del parser.
Ejecutar: python tests/generate_fixtures.py
"""
from pathlib import Path
from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"
FIXTURES_DIR.mkdir(exist_ok=True)


def create_podologia():
    """
    Estilo Podología: logo/título en filas 1-2, headers en fila 3, datos desde fila 4.
    Columnas en posiciones "raras": ID en B, Descripción en C, Estado en G, N° Serie en H.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "RELEVAMIENTO"

    # Filas institucionales (filas 1-2)
    ws.cell(row=1, column=1, value="HOSPITAL MELCHOR ROMERO")
    ws.cell(row=2, column=1, value="Relevamiento de Bienes - Podología y Peluquería")

    # Headers en fila 3 — con nombres no estándar
    headers = {
        1: "N° ITEM",  # columna no reconocida (no es "identificacion")
        2: "N° DE IDENTIDAD",  # → numero_id
        3: "DESCRIPCION DEL BIEN",  # → descripcion
        4: "CANTIDAD",  # → cantidad
        5: "ORDEN DE COMPRA",  # → numero_compra
        6: "N° DE EXPEDIENTE DE ORDEN DE COMPRA",  # → expediente_oc
        7: "ESTADO (ACTIVO/BAJA/DEFINITIVA /BAJA POR TRASLADO",  # → estado
        8: "N° DE SERIE",  # → numero_serie
        9: "UBICACION",  # → servicios
        10: "SIEM",  # → siem
        11: "OBSERVACIONES",  # → observaciones
    }
    for col, header in headers.items():
        ws.cell(row=3, column=col, value=header)

    # Datos desde fila 4
    data = [
        [1, "POD-001", "Sillón podológico eléctrico", 1, "OC-123/2024", "EX-001/2024", "ACTIVO", "SN-001", "Podología Y Peluquería", "Si", "Buen estado"],
        [2, "POD-002", "Instrumental de podología", 3, "NO", "NO", "ACTIVO", "NO", "Podología Y Peluquería", "No", ""],
        [3, "POD-003", "Esterilizador de calor seco", 1, "OC-456/2024", "EX-002/2024", "MANTENIMIENTO", "SN-003", "Podología Y Peluquería", "Si", "En reparación"],
        [4, "POD-004", "Lámpara de pie articulada", 2, "NO", "NO", "BAJA", "SN-004", "Podología Y Peluquería", "No", "Dañada"],
    ]
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    path = FIXTURES_DIR / "relevamiento_podologia.xlsx"
    wb.save(str(path))
    print(f"  ✅ {path.name}")


def create_neonatologia():
    """
    Estilo Neonatología: headers en fila 1, datos desde fila 2.
    Columnas en posiciones DISTINTAS a podología.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Neonatología"

    # Headers en fila 1 — nombres diferentes
    headers = {
        1: "IDENTIFICACION",  # → numero_id
        2: "DESCRIPCION",  # → descripcion
        3: "CANT",  # → cantidad
        4: "N° DE EXPEDIENTE POR DONACION",  # → expediente_donacion
        5: "N° DE EXPEDIENTE\nPOR ORDEN DE COMPRA",  # → expediente_oc (con newline!)
        6: "ESTADO DEL BIEN",  # → estado
        7: "UBICACION",  # → servicios
        8: "OBSERVACIONES",  # → observaciones
        9: "NUMERO DE SERIE",  # → numero_serie
        10: "N SIEM",  # → siem
    }
    for col, header in headers.items():
        ws.cell(row=1, column=col, value=header)

    # Datos
    data = [
        ["NEO-001", "Monitor de signos vitales neonatal", 2, "NO", "EX-NEO-001", "ACTIVO", "Pediatría Y Neonatología", "Funcional", "SN-NEO-001", "Si"],
        ["NEO-002", "Incubadora Dräger", 1, "EX-DON-001", "NO", "ACTIVO", "Pediatría Y Neonatología", "", "SN-NEO-002", "Si"],
        ["NEO-003", "Bomba de infusión", 4, "NO", "EX-NEO-002", "ACTIVO", "Pediatría Y Neonatología", "Nuevas", "NO", "No"],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    path = FIXTURES_DIR / "relevamiento_neonatologia.xlsx"
    wb.save(str(path))
    print(f"  ✅ {path.name}")


def create_modelo_limpio():
    """
    Excel "limpio" usando los encabezados canónicos del modelo descargable.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    headers = [
        "IDENTIFICACION", "DESCRIPCION DEL BIEN", "CANTIDAD", "ESTADO",
        "N° DE SERIE", "ORDEN DE COMPRA",
        "N° DE EXPEDIENTE POR ORDEN DE COMPRA",
        "N° DE EXPEDIENTE POR DONACION",
        "UBICACION", "SIEM", "OBSERVACIONES",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    data = [
        ["LIM-001", "Escritorio de oficina", 1, "ACTIVO", "SN-LIM-001", "OC-100", "EX-100", "NO", "Informatica", "No", ""],
        ["LIM-002", "Silla ergonómica", 5, "ACTIVO", "NO", "NO", "NO", "NO", "Informatica", "No", "Lote nuevo"],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    path = FIXTURES_DIR / "modelo_limpio.xlsx"
    wb.save(str(path))
    print(f"  ✅ {path.name}")


def create_columna_inventada():
    """
    Excel con una columna no reconocida: "CHAMUYO".
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    headers = [
        "IDENTIFICACION", "DESCRIPCION DEL BIEN", "CANTIDAD", "ESTADO",
        "CHAMUYO",  # ← columna inventada
        "UBICACION",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    ws.cell(row=2, column=1, value="INV-001")
    ws.cell(row=2, column=2, value="Objeto misterioso")
    ws.cell(row=2, column=3, value=1)
    ws.cell(row=2, column=4, value="ACTIVO")
    ws.cell(row=2, column=5, value="dato chamuyo")
    ws.cell(row=2, column=6, value="Deposito General")

    path = FIXTURES_DIR / "con_columna_inventada.xlsx"
    wb.save(str(path))
    print(f"  ✅ {path.name}")


def create_vacio():
    """Excel completamente vacío."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacío"
    path = FIXTURES_DIR / "vacio.xlsx"
    wb.save(str(path))
    print(f"  ✅ {path.name}")


def create_valores_no():
    """
    Excel con valores "NO" en campos opcionales y cantidad no numérica.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    headers = [
        "IDENTIFICACION", "DESCRIPCION DEL BIEN", "CANTIDAD",
        "ESTADO", "N° DE SERIE", "UBICACION", "OBSERVACIONES",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Fila con "NO" en campos opcionales
    ws.cell(row=2, column=1, value="VAL-001")
    ws.cell(row=2, column=2, value="Mesa de examen")
    ws.cell(row=2, column=3, value="tres")  # cantidad no numérica
    ws.cell(row=2, column=4, value="ACTIVO")
    ws.cell(row=2, column=5, value="NO")  # serie = NO → vacío
    ws.cell(row=2, column=6, value="NO")  # ubicación = NO → vacío
    ws.cell(row=2, column=7, value="NO")  # observaciones = NO → vacío

    path = FIXTURES_DIR / "valores_no.xlsx"
    wb.save(str(path))
    print(f"  ✅ {path.name}")


if __name__ == "__main__":
    print("Generando fixtures de test...")
    create_podologia()
    create_neonatologia()
    create_modelo_limpio()
    create_columna_inventada()
    create_vacio()
    create_valores_no()
    print("✅ Todas las fixtures generadas.")
