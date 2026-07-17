# core/parsers.py
"""
Parser inteligente de relevamientos Excel.

Detecta automáticamente la hoja con datos, la fila de encabezados,
y mapea columnas por sinónimos para soportar planillas con formatos
distintos (podología, neonatología, etc.) sin que el usuario tenga
que modificar el archivo.

Para agregar un sinónimo nuevo basta un commit de una línea
añadiendo el string al array correspondiente en HEADER_SYNONYMS.
"""
import re
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from unidecode import unidecode


# ──────────────────────────────────────────────
# Excepciones propias
# ──────────────────────────────────────────────

class ParseError(Exception):
    """Error base de parseo."""


class EmptyWorkbookError(ParseError):
    """El libro no tiene hojas con datos."""


class HeaderNotFoundError(ParseError):
    """No se detectó una fila de encabezados válida."""


# ──────────────────────────────────────────────
# Diccionario de sinónimos
# ──────────────────────────────────────────────

HEADER_SYNONYMS: dict[str, list[str]] = {
    'numero_id': [
        'identificacion', 'n de identidad', 'identidad',
        'n id', 'numero de identidad', 'codigo',
        'numero de identificacion', 'id', 'id patrimonial',
        'n de identificacion',
    ],
    'descripcion': [
        'descripcion del bien', 'descripcion', 'bien',
        'detalle del bien', 'detalle', 'nombre',
    ],
    'cantidad': ['cantidad', 'cant', 'qty'],
    'estado': [
        'estado', 'estado del bien',
        'estado activo baja definitiva baja por traslado',
        'estado activobajadefinitivabaja por traslado',
        'condicion',
    ],
    'numero_serie': ['n de serie', 'numero de serie', 'serie', 'nro serie'],
    'numero_compra': [
        'orden compra', 'orden de compra', 'oc',
        'n de orden de compra', 'n compra', 'numero compra',
        'nro compra',
    ],
    'expediente_oc': [
        'n de expediente de orden de compra',
        'n de expediente de oeden de compra',  # typo real en planillas
        'n de expediente por orden de compra',  # variante con "por"
        'expediente oc', 'expediente orden compra',
        'n expediente', 'numero expediente', 'expediente',
        'exp', 'nro exp', 'n de expediente',
    ],
    'expediente_donacion': [
        'n de expediente por donacion',
        'expediente donacion', 'donacion',
    ],
    'servicios': [
        'ubicacion', 'servicio', 'servicios', 'sector', 'area',
    ],
    'siem': ['siem', 'sien', 'n siem', 'estado siem'],
    'observaciones': ['observaciones', 'observacion', 'obs', 'notas', 'comentarios'],
    'cuenta_codigo': ['cuenta codigo', 'cuenta', 'cod cuenta', 'codigo cuenta'],
    'nomenclatura': ['nomenclatura', 'nomenclatura bienes', 'cod nomenclatura'],
}

# Índice invertido: sinónimo normalizado → campo canónico
_SYNONYM_INDEX: dict[str, str] = {}
for _field, _synonyms in HEADER_SYNONYMS.items():
    for _syn in _synonyms:
        _SYNONYM_INDEX[_syn] = _field


# ──────────────────────────────────────────────
# Funciones de parseo
# ──────────────────────────────────────────────

def normalize_header(text: str) -> str:
    """
    Normaliza un texto de encabezado para matching por sinónimos.

    - Lowercase
    - Strip accents (con unidecode)
    - Reemplazar °/º por ''
    - Reemplazar \\n por espacio
    - Eliminar caracteres no alfanuméricos (excepto espacio)
    - Colapsar espacios múltiples
    - Strip

    Ejemplo:
        'N° DE EXPEDIENTE\\nPOR DONACIÓN  ' → 'n de expediente por donacion'
    """
    if not text:
        return ''
    text = str(text)
    # Reemplazar grado/ordinal
    text = text.replace('°', '').replace('º', '')
    # Newlines → espacio
    text = text.replace('\n', ' ').replace('\r', ' ')
    # Strip accents
    text = unidecode(text)
    # Lowercase
    text = text.lower()
    # Solo alfanuméricos y espacios
    text = re.sub(r'[^a-z0-9 ]', ' ', text)
    # Colapsar espacios
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def detect_data_sheet(workbook) -> Worksheet:
    """
    Selecciona la hoja con datos del workbook.

    - Si hay una sola hoja con datos, la devuelve.
    - Si hay varias, elige la que tiene más celdas no vacías.
    - Si todas están vacías, lanza EmptyWorkbookError.
    """
    if not workbook.sheetnames:
        raise EmptyWorkbookError("El libro no tiene hojas.")

    best_sheet = None
    best_count = 0

    for name in workbook.sheetnames:
        ws = workbook[name]
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 500), values_only=True):
            for cell in row:
                if cell is not None and str(cell).strip():
                    count += 1
        if count > best_count:
            best_count = count
            best_sheet = ws

    if best_count == 0:
        raise EmptyWorkbookError("Todas las hojas del libro están vacías.")

    return best_sheet


def detect_header_row(worksheet: Worksheet, max_scan: int = 15) -> int:
    """
    Detecta la fila de encabezados escaneando las primeras `max_scan` filas.

    La primera fila donde al menos 3 celdas normalizadas matcheen algún
    valor de HEADER_SYNONYMS es la fila de headers.

    Returns:
        Número de fila (1-based) de los headers.

    Raises:
        HeaderNotFoundError si no se detecta ninguna.
    """
    for row_num in range(1, max_scan + 1):
        matches = 0
        for cell in worksheet[row_num]:
            if cell.value is not None:
                normalized = normalize_header(str(cell.value))
                if normalized and normalized in _SYNONYM_INDEX:
                    matches += 1
        if matches >= 3:
            return row_num

    raise HeaderNotFoundError(
        f"No se detectó una fila de encabezados válida en las primeras {max_scan} filas. "
        "Verificá que el archivo tenga columnas como IDENTIFICACION, DESCRIPCION, ESTADO, etc."
    )


def map_columns(worksheet: Worksheet, header_row: int) -> tuple[dict[str, int], list[str]]:
    """
    Mapea columnas del worksheet a campos canónicos usando el diccionario de sinónimos.

    Args:
        worksheet: Hoja de cálculo.
        header_row: Número de fila (1-based) con los encabezados.

    Returns:
        Tupla (mapping, unrecognized):
        - mapping: {campo_canonico: indice_columna_1based}
        - unrecognized: lista de textos de columnas no reconocidas.
    """
    mapping: dict[str, int] = {}
    unrecognized: list[str] = []

    for cell in worksheet[header_row]:
        if cell.value is None:
            continue
        raw_text = str(cell.value).strip()
        if not raw_text:
            continue

        normalized = normalize_header(raw_text)
        if not normalized:
            continue

        canonical = _SYNONYM_INDEX.get(normalized)
        if canonical:
            # Solo tomar la primera columna que matchea un campo canónico
            if canonical not in mapping:
                mapping[canonical] = cell.column
        else:
            unrecognized.append(raw_text)

    return mapping, unrecognized


# ──────────────────────────────────────────────
# Resultado del parseo
# ──────────────────────────────────────────────

@dataclass
class ParseResult:
    """Resultado del parseo de un relevamiento Excel."""
    rows: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    meta: dict = field(default_factory=dict)


# ──────────────────────────────────────────────
# Función principal
# ──────────────────────────────────────────────

def _cell_value(worksheet: Worksheet, row: int, col: int) -> Optional[str]:
    """Obtiene el valor de una celda como string limpio, o None si está vacía."""
    val = worksheet.cell(row=row, column=col).value
    if val is None:
        return None
    text = str(val).strip()
    if text.lower() in ('', 'nan', 'none'):
        return None
    return text


def _is_no(value: Optional[str]) -> bool:
    """Retorna True si el valor es el literal 'NO' (convención de planillas viejas)."""
    return value is not None and value.strip().upper() == 'NO'


def _clean_optional(value: Optional[str]) -> str:
    """Limpia un campo opcional: si es 'NO' o vacío, devuelve ''."""
    if value is None or _is_no(value):
        return ''
    return value


def _parse_cantidad(value: Optional[str], row_num: int, warnings: list[str]) -> int:
    """Parsea cantidad: si no es numérico, devuelve 1 y agrega warning."""
    if value is None or _is_no(value):
        return 1
    try:
        return max(int(float(value)), 1)
    except (ValueError, TypeError):
        warnings.append(f"Fila {row_num}: cantidad '{value}' no es numérica, se usó 1.")
        return 1


def _map_estado(value: Optional[str]) -> str:
    """Mapea un valor de estado del Excel a los ESTADO_CHOICES del modelo."""
    if not value or _is_no(value):
        return ''
    t = value.lower()
    if 'manten' in t:
        return 'MANTENIMIENTO'
    if 'baja' in t:
        return 'BAJA'
    if 'inac' in t:
        return 'INACTIVO'
    if 'activ' in t:
        return 'ACTIVO'
    return ''


def parse_relevamiento(file_path: str, default_year: int = None) -> ParseResult:
    """
    Parsea un archivo de relevamiento Excel.

    Orquesta: detect_data_sheet → detect_header_row → map_columns → iterar filas.

    Args:
        file_path: Ruta al archivo .xlsx.
        default_year: Año para fecha_alta si no se proporciona.

    Returns:
        ParseResult con las filas parseadas, warnings y metadata.

    Raises:
        EmptyWorkbookError: Si el libro no tiene datos.
        HeaderNotFoundError: Si no se detectan encabezados.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = detect_data_sheet(wb)
        header_row = detect_header_row(ws)
        mapping, unrecognized = map_columns(ws, header_row)

        result = ParseResult(
            meta={
                'hoja': ws.title,
                'header_row': header_row,
                'columnas_mapeadas': {k: v for k, v in mapping.items()},
                'columnas_no_reconocidas': unrecognized,
            }
        )

        # Funciones helper para obtener valores
        def get_field(row_num: int, field_name: str) -> Optional[str]:
            col = mapping.get(field_name)
            if col is None:
                return None
            return _cell_value(ws, row_num, col)

        # Determinar fecha_alta default
        if default_year:
            default_date = date(default_year, 1, 1)
        else:
            default_date = date.today()

        # Iterar filas de datos (desde header_row + 1 hasta el final)
        max_row = ws.max_row or header_row
        for row_num in range(header_row + 1, max_row + 1):
            # Obtener descripcion y numero_id primero para decidir si saltar
            descripcion = get_field(row_num, 'descripcion')
            numero_id = get_field(row_num, 'numero_id')

            # Verificar si la fila está completamente vacía
            all_empty = True
            for col_idx in mapping.values():
                val = _cell_value(ws, row_num, col_idx)
                if val is not None:
                    all_empty = False
                    break
            if all_empty:
                continue

            # Saltar filas donde tanto descripcion como numero_id están vacías
            if not descripcion and not numero_id:
                continue

            # Parsear campos
            cantidad_raw = get_field(row_num, 'cantidad')
            cantidad = _parse_cantidad(cantidad_raw, row_num, result.warnings)

            estado_raw = get_field(row_num, 'estado')
            estado = _map_estado(estado_raw)

            numero_serie = _clean_optional(get_field(row_num, 'numero_serie'))
            numero_compra = _clean_optional(get_field(row_num, 'numero_compra'))
            observaciones = _clean_optional(get_field(row_num, 'observaciones'))
            servicios = _clean_optional(get_field(row_num, 'servicios'))
            siem_raw = (get_field(row_num, 'siem') or '').strip().lower()
            siem = 'Si' if siem_raw and siem_raw not in ('no', '0', 'false', 'n', '') else 'No'
            cuenta_codigo = _clean_optional(get_field(row_num, 'cuenta_codigo'))
            nomenclatura = _clean_optional(get_field(row_num, 'nomenclatura'))

            # Expediente: prioridad expediente_oc > expediente_donacion
            exp_oc = get_field(row_num, 'expediente_oc')
            exp_don = get_field(row_num, 'expediente_donacion')
            expediente_str = ''
            if exp_oc and not _is_no(exp_oc):
                expediente_str = exp_oc
            elif exp_don and not _is_no(exp_don):
                expediente_str = exp_don

            # Limpiar numero_id
            numero_id_clean = None
            if numero_id:
                nid = numero_id.strip()
                if nid.upper() not in ('NO', '-', ''):
                    numero_id_clean = nid[:50]

            # Nombre: el modelo BienPatrimonial usa campo 'nombre' (max 200)
            nombre = (descripcion or numero_serie or 'SIN NOMBRE')[:200]

            row_data = {
                'nombre': nombre,
                'descripcion': descripcion or '',
                'cantidad': cantidad,
                'numero_identificacion': numero_id_clean,
                'numero_serie': numero_serie[:100] if numero_serie else '',
                'numero_compra': numero_compra[:50] if numero_compra else '',
                'estado': estado or 'ACTIVO',
                'servicios': servicios[:200] if servicios else '',
                'observaciones': observaciones or '',
                'siem': siem,
                'cuenta_codigo': cuenta_codigo[:20] if cuenta_codigo else '',
                'nomenclatura_bienes': nomenclatura[:200] if nomenclatura else '',
                # origen y precio: siempre vacíos desde Excel
                'origen': '',
                'valor_adquisicion': None,
                # fecha_alta
                'fecha_adquisicion': default_date,
                # expediente: se resuelve en la vista (se necesita FK)
                '_expediente_str': expediente_str[:50] if expediente_str else '',
            }

            result.rows.append(row_data)

        return result
    finally:
        wb.close()
