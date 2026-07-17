#!/usr/bin/env python3
"""
Genera el modelo de referencia para carga masiva.
Ejecutar una sola vez:
    python core/plantillas/generar_modelo.py
"""
import sys
from pathlib import Path

# Agregar raíz del proyecto al path
root = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(root))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()

# ── Hoja 1: Datos ──
ws_datos = wb.active
ws_datos.title = "Datos"

headers = [
    "IDENTIFICACION",
    "DESCRIPCION DEL BIEN",
    "CANTIDAD",
    "ESTADO",
    "N° DE SERIE",
    "ORDEN DE COMPRA",
    "N° DE EXPEDIENTE POR ORDEN DE COMPRA",
    "N° DE EXPEDIENTE POR DONACION",
    "UBICACION",
    "SIEM",
    "OBSERVACIONES",
]

# Estilos
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="00838F", end_color="00838F", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col_idx, header in enumerate(headers, 1):
    cell = ws_datos.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Anchos de columna razonables
widths = [18, 35, 12, 18, 18, 18, 38, 38, 22, 10, 30]
for col_idx, w in enumerate(widths, 1):
    ws_datos.column_dimensions[ws_datos.cell(row=1, column=col_idx).column_letter].width = w

# Freezar la primera fila
ws_datos.freeze_panes = "A2"

# ── Hoja 2: INSTRUCCIONES ──
ws_instr = wb.create_sheet("INSTRUCCIONES")
ws_instr.sheet_properties.tabColor = "FFC107"

instrucciones = [
    "INSTRUCCIONES PARA CARGA MASIVA",
    "",
    "Este modelo es solo una referencia para relevamientos nuevos.",
    "El sistema acepta archivos con otros nombres de columna.",
    "",
    "Ejemplos de nombres alternativos aceptados:",
    '  • "N° DE IDENTIDAD" en lugar de "IDENTIFICACION"',
    '  • "ESTADO DEL BIEN" en lugar de "ESTADO"',
    '  • "DESCRIPCION" en lugar de "DESCRIPCION DEL BIEN"',
    '  • "SERVICIO" o "SECTOR" en lugar de "UBICACION"',
    "",
    "Si tu relevamiento ya está armado, podés subirlo tal cual.",
    "",
    "Los campos ORIGEN, PRECIO y FECHA DE CARGA se completan",
    "después de subir el archivo, en la pantalla del sistema.",
    "No los pongas en el Excel.",
    "",
    "Formato aceptado: .xlsx o .xls",
    "Tamaño máximo: 10 MB",
]

title_font = Font(name="Calibri", bold=True, size=14, color="00838F")
body_font = Font(name="Calibri", size=11)

for row_idx, text in enumerate(instrucciones, 1):
    cell = ws_instr.cell(row=row_idx, column=1, value=text)
    if row_idx == 1:
        cell.font = title_font
    else:
        cell.font = body_font

ws_instr.column_dimensions["A"].width = 60

# Guardar
output_path = Path(__file__).resolve().parent / "modelo_carga_masiva.xlsx"
wb.save(str(output_path))
print(f"✅ Modelo guardado en: {output_path}")
